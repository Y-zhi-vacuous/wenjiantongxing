"""学生账号 xlsx 批量导入导出"""
import io
from urllib.parse import quote
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook, load_workbook

from app.db import get_db
from app.models import User, UserRole, Class, ClassStudent
from app.auth import get_current_user, hash_password

router = APIRouter(prefix="/classes", tags=["学生管理"])

TEMPLATE_HEADERS = ["班级", "学生姓名", "学号/账号", "密码"]
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _make_xlsx_response(wb: Workbook, filename: str) -> Response:
    output = io.BytesIO()
    wb.save(output)
    content = output.getvalue()
    encoded = quote(filename)
    return Response(
        content=content,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.get("/{class_id}/students/template")
async def download_template(
    class_id: int,
    teacher: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if teacher.role.value != "teacher":
        raise HTTPException(status_code=403, detail="仅教师可操作")
    cls = await db.get(Class, class_id)
    if not cls:
        raise HTTPException(status_code=404, detail="班级不存在")

    wb = Workbook()
    ws = wb.active
    ws.title = "学生导入模板"
    ws.append(TEMPLATE_HEADERS)
    ws.append([cls.name, "张小明", "zhangxm", "abc123"])
    ws.append([cls.name, "李小华", "lixh", "abc123"])
    for col, width in zip("ABCD", [20, 16, 18, 16]):
        ws.column_dimensions[col].width = width

    return _make_xlsx_response(wb, f"student_template_{cls.name}.xlsx")


@router.post("/{class_id}/students/import")
async def import_students(
    class_id: int,
    file: UploadFile = File(...),
    teacher: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if teacher.role.value != "teacher":
        raise HTTPException(status_code=403, detail="仅教师可操作")
    cls = await db.get(Class, class_id)
    if not cls:
        raise HTTPException(status_code=404, detail="班级不存在")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="无法读取 xlsx 文件，请使用下载的模板格式")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        display_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        username = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        password = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        if not username or not password or not display_name:
            skipped += 1
            continue

        existing = await db.execute(select(User).where(User.username == username))
        if existing.scalar_one_or_none():
            skipped += 1
            continue

        student = User(
            username=username, password_hash=hash_password(password),
            default_password=password, role=UserRole.student,
            display_name=display_name, school=teacher.school, grade="九年级",
        )
        db.add(student)
        await db.flush()
        db.add(ClassStudent(class_id=class_id, student_id=student.id))
        created += 1

    await db.commit()
    return {"message": f"导入完成：成功 {created} 人，跳过 {skipped} 人", "created": created, "skipped": skipped}


@router.get("/{class_id}/students/export")
async def export_students(
    class_id: int,
    teacher: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if teacher.role.value != "teacher":
        raise HTTPException(status_code=403, detail="仅教师可操作")
    cls = await db.get(Class, class_id)
    if not cls:
        raise HTTPException(status_code=404, detail="班级不存在")

    result = await db.execute(
        select(User).join(ClassStudent).where(ClassStudent.class_id == class_id)
    )
    students = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "学生列表"
    ws.append(["班级", "学生姓名", "学号/账号", "当前密码", "作文数"])

    for user in students:
        from sqlalchemy import func
        from app.models import Essay
        essay_count = await db.scalar(
            select(func.count(Essay.id)).where(Essay.student_id == user.id)
        )
        password = user.default_password or "(已修改)"
        ws.append([cls.name, user.display_name, user.username, password, essay_count or 0])

    for col, width in zip("ABCDE", [20, 16, 18, 16, 10]):
        ws.column_dimensions[col].width = width

    return _make_xlsx_response(wb, f"students_{cls.name}.xlsx")
